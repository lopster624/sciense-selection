import datetime
import re
from io import BytesIO
from openpyxl import load_workbook, Workbook
from openpyxl.utils.cell import get_column_letter

from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q
from django.db.models.functions import Lower
from django.shortcuts import get_object_or_404, redirect
from docxtpl import DocxTemplate

from account.models import Member, Affiliation, Booking, Role, BookingType
from engine.middleware import logger
from utils.calculations import get_current_draft_year, convert_float
from utils.constants import BOOKED, MEANING_COEFFICIENTS, PATH_TO_RATING_LIST, \
    PATH_TO_CANDIDATES_LIST, PATH_TO_EVALUATION_STATEMENT
from utils import constants as const
from utils.exceptions import MaxAffiliationBookingException
from utils.calculations import convert_date_str_to_datetime, convert_datetime_to_str, convert_phone_format

from .models import Application, AdditionField, AdditionFieldApp, MilitaryCommissariat, Education, Universities, \
    Specialization, Competence, ApplicationCompetencies, Direction


def check_role(user, role_name):
    """ Проверяет, совпадает ли роль пользователя с заданной через параметр role_name """
    member = Member.objects.select_related('role').get(user=user)
    if member.role.role_name == role_name:
        return True
    return False


def check_affiliation_max_count(pk, booking_type, affiliation):
    """
    Проверяет, что в affiliation забронировано в сезоне призыва заявки(pk) не больше макс. количества.
    :param pk: идентификатор анкеты
    :param booking_type: тип бронирования
    :param affiliation: объект affiliation
    :return: True, если максимальное количество не превышено
    """
    slave_draft_info = Application.objects.filter(pk=pk).values('draft_year', 'draft_season')
    if Booking.objects.filter(booking_type=booking_type, affiliation=affiliation,
                              slave__application__draft_year=slave_draft_info[0].get('draft_year'),
                              slave__application__draft_season=slave_draft_info[0].get(
                                  'draft_season')).count() + 1 > const.MAX_BOOKED_ON_AFFILIATION:
        raise MaxAffiliationBookingException(
            f'Невозможно отобрать заявку в {affiliation.platoon} взвод {affiliation.company} роты. '
            f'Превышено максимальное количество отобранных заявок ({const.MAX_BOOKED_ON_AFFILIATION} шт.)'
            f' в данный взвод в этом призыве! Вам необоходимо сначало удалить из отбора другую заявку,'
            f' чтобы добавить эту.')
    return True


def check_permission_decorator(role_name=None):
    """ Кидает исключение PermissionDenied, если роль user!=role_name
    или текущий пользователь не является пользователем с переданным pk"""

    def decorator(func):
        def wrapper(self, request, pk, *args, **kwargs):
            if request.user.member.role.role_name == role_name:
                return func(self, request, pk, *args, **kwargs)
            member = Member.objects.filter(application__id=pk).first()
            if member != request.user.member:
                raise PermissionDenied('Недостаточно прав для входа на данную страницу.')
            return func(self, request, pk, *args, **kwargs)

        return wrapper

    return decorator


def check_final_decorator(func):
    """Декоратор, который проверяет, что анкету пользователя можно редактировать.
     В противном случае редиректит на предыдущую страницу"""

    def wrapper(self, request, pk, *args, **kwargs):
        user_app = get_object_or_404(Application.objects.only('is_final'), pk=pk)
        if user_app.is_final:
            return redirect(request.path_info)
        return func(self, request, pk, *args, **kwargs)

    return wrapper


class WordTemplate:
    """ Класс для создания шаблона ворд документа по файлу, который через путь - path_to_template """

    def __init__(self, request, path_to_template):
        self.request = request
        self.path = path_to_template

    def create_word_in_buffer(self, context):
        """ Создает ворд документ и добавлет в него данные и сохраняет в буфер """
        template = DocxTemplate(self.path)
        user_docx = BytesIO()
        template.render(context=context)
        template.save(user_docx)
        user_docx.seek(0)
        return user_docx

    def create_context_to_interview_list(self, pk):
        """ Создает контекст для шаблона - 'Лист собеседования' """
        user_app = Application.objects.select_related('member').prefetch_related('education').defer('id').get(pk=pk)
        user_education = user_app.education.order_by('-end_year').values()
        user_education = user_education[0] if user_education else {}
        context = {**user_app.__dict__, **user_education}
        context.update({'father_name': user_app.member.father_name, 'phone': user_app.member.phone})
        context.update({'first_name': user_app.member.user.first_name, 'last_name': user_app.member.user.last_name})
        return context

    def create_context_to_word_files(self, document_type, all_directions=None):
        """ Создает контексты для шаблонов итоговых документов """
        current_year, current_season = get_current_draft_year()
        fixed_directions = self.request.user.member.affiliations.select_related('direction').all() \
            if not all_directions else Affiliation.objects.select_related('direction').all()

        context = {'directions': []}
        for direction in fixed_directions:
            platoon_data = {
                'name': direction.direction.name,
                'company_number': direction.company,
                'members': []
            }
            booked = Booking.objects.select_related('slave').filter(affiliation=direction, booking_type__name=BOOKED)
            booked_slaves = [b.slave for b in booked]
            booked_user_apps = Application.objects.select_related('scores', 'member__user').prefetch_related(
                'education'). \
                filter(member__in=booked_slaves, draft_year=current_year, draft_season=current_season[0]).all()

            for i, user_app in enumerate(booked_user_apps):
                user_last_education = user_app.education.all()[0]
                general_info, additional_info = {'number': i + 1,
                                                 'first_name': user_app.member.user.first_name,
                                                 'last_name': user_app.member.user.last_name,
                                                 'father_name': user_app.member.father_name,
                                                 'final_score': convert_float(user_app.final_score),
                                                 }, {}
                if document_type == PATH_TO_CANDIDATES_LIST:
                    additional_info = self._get_candidates_info(user_app, user_last_education)
                elif document_type == PATH_TO_RATING_LIST:
                    additional_info = self._get_rating_info(user_app, user_last_education)
                elif document_type == PATH_TO_EVALUATION_STATEMENT:
                    additional_info = self._get_evaluation_st_info(user_app)
                platoon_data['members'].append({**additional_info, **general_info})
            if platoon_data['members']:
                context['directions'].append(platoon_data)
        return context

    def create_context_to_psychological_test(self, user_test_result, questions, user_answers):
        """ Создает контекст для шаблона - 'Психологического теста ОПВС - 2' """
        user_app = Application.objects.only('birth_day').get(member=user_test_result.member)
        member = user_test_result.member
        context = {
            'full_name': f'{member.user.last_name} {member.user.first_name} {member.father_name}',
            'b_day': user_app.birth_day.strftime('%d %m %Y'),
            'now': datetime.datetime.now().strftime('%d %m %Y'),
            'position': 'гражданин',
            'questions': []
        }
        for i, question in enumerate(questions, 1):
            context['questions'].append({
                'num': i,
                'answers': [ans.id for ans in question.answer_options.all()],
                'response': user_answers.get(question.id)
            })
        return context

    def _get_evaluation_st_info(self, user_app):
        return {
            **{k: convert_float(v) for k, v in MEANING_COEFFICIENTS.items()},
            **{k: convert_float(v) for k, v in user_app.scores.__dict__.items() if isinstance(v, float)},
        }

    def _get_rating_info(self, user_app, user_last_education):
        return {
            'birth_day': user_app.birth_day,
            'military_commissariat': user_app.military_commissariat,
            'university': user_last_education.university,
            'specialization': user_last_education.specialization,
            'avg_score': convert_float(user_last_education.avg_score),
        }

    def _get_candidates_info(self, user_app, user_last_education):
        commissariat = MilitaryCommissariat.objects.filter(name=user_app.military_commissariat).first()
        return {
            'subject': commissariat.subject if commissariat else '',
            'birth_day': user_app.birth_day.year,
            'avg_score': convert_float(user_last_education.avg_score),
        }


def check_booking_our_or_exception(pk, user):
    """Проверяет, что пользователь с айди анкеты pk был забронирован на направления пользователя user и
    рейзит ошибку, если не забронирован."""
    if not is_booked_our(pk, user):
        raise PermissionDenied('Данный пользователь не отобран на ваше направление.')


def is_booked_our(pk, user):
    """
    Возвращает True, если пользователь с айди анкеты = pk забронирован на направления user,
    в обратном случае - False
    """
    app = get_object_or_404(Application, pk=pk)
    master_affiliations = Affiliation.objects.filter(member=user.member)
    return Booking.objects.filter(slave=app.member, booking_type__name=BOOKED,
                                  affiliation__in=master_affiliations).exists()


def add_additional_fields(request, user_app):
    additional_fields = [int(re.search('\d+', field).group(0)) for field in request.POST
                         if const.NAME_ADDITIONAL_FIELD_TEMPLATE in field]
    if additional_fields:
        addition_fields = AdditionField.objects.filter(pk__in=additional_fields)
        for field in addition_fields:
            AdditionFieldApp.objects.update_or_create(addition_field=field, application=user_app,
                                                      defaults={'value': request.POST.get(
                                                          f"{const.NAME_ADDITIONAL_FIELD_TEMPLATE}{field.id}")})


def get_additional_fields(request):
    additional_fields = {int(re.search('\d+', field).group(0)): request.POST.get(field) for field in request.POST
                         if const.NAME_ADDITIONAL_FIELD_TEMPLATE in field}
    return additional_fields


def get_cleared_query_string_of_page(query_string):
    """Возвращает строку query-параметров без параметра page"""
    if not query_string:
        return ''
    return '&'.join(param for param in query_string.split('&') if 'page=' not in param) + '&'


def get_sorted_queryset(apps, ordering):
    """Возвращает отсортированый queryset по our_direction и ordering(если есть)"""
    if ordering:
        if ordering in ['subject_name', 'birth_place']:
            # если поля для поиска текстовое, то сортируем в lowercase
            ordering = Lower(ordering)
        return apps.order_by('-our_direction', ordering)
    return apps.order_by('-our_direction', '-create_date')


def get_form_data(get_dict):
    """Если в словаре get_dict содержится что-то кроме page, то возвращает его. В противном случае возвращает None"""
    orig_dict = dict(get_dict)
    orig_dict.pop('page', None)
    if not orig_dict:
        return None
    return get_dict


class Questionnaires:
    def __init__(self, file):
        wb = load_workbook(file, read_only=True)
        sheet = wb.active
        self.sheet = sheet

    def create_member(self, user_params):
        new_user = self._create_new_user(user_params)
        return Member.objects.create(user=new_user, role=Role.objects.get(role_name=const.SLAVE_ROLE_NAME),
                                     father_name=user_params.get('father_name'), phone=user_params.get('phone'))

    def add_applications_to_db(self):
        result = {
            'errors': [],
            'accepted': [],
        }
        for line_number, row in enumerate(self.sheet.iter_rows(min_row=2, values_only=True), start=1):
            params = {name: cell for name, cell in zip(const.EXCEL_TABLE_HEADERS_FOR_IMPORT_APPS, row)}
            if not params['id']:
                continue
            user_params = self._get_params_for_member_create(params)
            if member := self._if_member_exists(user_params):
                if not self._delete_app_if_not_booked(member):
                    result['errors'].append(
                        f"Анкета пользователя {params['full_name']} (id: {params['id']}) уже была создана и забронированна ранее!")
                    continue
                result['errors'].append(
                    f"Анкета пользователя {params['full_name']} (id: {params['id']}) уже была создана и была обновлена!")

            try:
                with transaction.atomic():
                    member = self.create_member(user_params) if not member else member
                    new_app = self.create_application(member, params)
                    self.add_education(new_app, params)
                    self.add_additional_values(line_number, new_app, params)
                    new_app.update_scores()
                result['accepted'].append(
                    f"Анкета пользователя {params['full_name']} с id: {params['id']} успешно создана!")
            except Exception as e:
                result['errors'].append(f"Ошибка при создании анкеты с id:{params['id']} - {e}")
                logger.error(f'Ошибка в анкете с id: {params["id"]} - {e}')
        return result

    def _if_member_exists(self, user_params):
        return Member.objects.filter(Q(phone=user_params['phone']) | Q(user__email=user_params['email'])).first()

    def _delete_app_if_not_booked(self, member):
        if not Booking.objects.filter(slave=member, booking_type=BookingType.objects.get(name=BOOKED)):
            Application.objects.filter(member=member).delete()
            return True
        return False

    def create_application(self, member, params):
        birth_day = convert_date_str_to_datetime(params['birth_day'].split()[0], '%d.%m.%Y')
        ready_to_secret = self._convert_ready_to_secret(params['ready_to_secret'])
        draft_season = Converter.convert_draft_season_by_name(params['draft_season'][1:-1])
        return Application.objects.create(member=member, birth_day=birth_day, birth_place=params['birth_place'],
                                          nationality=params['nationality'],
                                          military_commissariat=params['military_commissariat'],
                                          group_of_health=params['group_of_health'],
                                          draft_year=int(params['draft_year']),
                                          draft_season=draft_season, ready_to_secret=ready_to_secret, )

    def _convert_ready_to_secret(self, ready_to_secret):
        return True if ready_to_secret == '[Да]' else False

    def add_education(self, app, params):
        education_type = Converter.convert_education_type_by_name(params['education_type'][1:-1])
        university, specialization = self._get_university_and_specialization(params['university'],
                                                                             params['specialization'])
        return Education.objects.create(application=app, education_type=education_type, university=university,
                                        specialization=specialization, end_year=int(params['end_year']),
                                        avg_score=round(float(params['avg_score'].replace(',', '.')), 2),
                                        name_of_education_doc=params['name_of_education_doc'],
                                        theme_of_diploma=params['theme_of_diploma'])

    def _get_university_and_specialization(self, university, specialization):
        uni = Universities.get_by_name_or_leave(university)
        spec = Specialization.get_by_name_or_leave(specialization)
        return uni, spec

    def add_additional_values(self, line_number, app, params):
        user_info = self._get_user_directions_achievements_and_other_info(line_number)
        self.add_directions_to_app(app, user_info['user_directions'])
        self.add_achievements_to_app(app, user_info['user_achievements'])
        self.update_additional_app_params(app, user_info)
        self.add_competencies_to_app(app, params)

    def add_achievements_to_app(self, app, user_achievements):
        for achievement in user_achievements:
            setattr(app, const.CONVERTER_ACHIEVEMENTS_NAMES_TO_MODEL_FIELDS[achievement], True)
        app.save()

    def add_directions_to_app(self, app, user_directions):
        directions = Direction.objects.filter(name__in=user_directions)
        app.directions.add(*directions)

    def update_additional_app_params(self, app, user_info):
        app.scientific_achievements = '\n'.join(user_info['scientific_achievements']) or ''
        app.scholarships = '\n'.join(user_info['scholarships']) or ''
        app.candidate_exams = '\n'.join(user_info['candidate_exams']) or ''
        app.sporting_achievements = '\n'.join(user_info['sporting_achievements']) or ''
        app.other_information = '\n'.join(user_info['other_information']) or ''
        app.hobby = '\n'.join(user_info['hobby']) or ''
        app.save()

    def add_competencies_to_app(self, app, params):
        user_selected_competencies = [comp for comp in const.EXIST_COMPETENCIES_ON_SITE if params[comp] != '[0]']
        competencies = Competence.objects.filter(name__in=user_selected_competencies)
        competencies_with_levels = [ApplicationCompetencies(application=app, competence=competence,
                                                            level=int(params[competence.name][1:-1]))
                                    for competence in competencies]
        ApplicationCompetencies.objects.bulk_create(competencies_with_levels)

    def _get_user_directions_achievements_and_other_info(self, line_number):
        following_app = None
        for j, row in enumerate(self.sheet.iter_rows(min_row=line_number + 2, values_only=True)):
            if row[0]:
                following_app = j + line_number + 2  # номер итерации + номер текущей строки с данными + номер 1 строки с данными excel
                break

        user_info = {
            'user_directions': [],
            'user_achievements': [],
            'scientific_achievements': [],
            'scholarships': [],
            'candidate_exams': [],
            'sporting_achievements': [],
            'hobby': [],
            'other_information': [],
        }

        last_row_with_user_info = following_app if following_app else self.sheet.max_row + 1
        for row in range(line_number + 2 - 1, last_row_with_user_info):
            for i, field_name, in enumerate(user_info.keys()):
                cell = self.sheet.cell(row=row, column=23 + i).value
                user_info[field_name].append(cell.strip()) if cell else None

        user_info['user_directions'] = [_[1:-1] for _ in user_info['user_directions']]
        user_info['user_achievements'] = [_[1:-1] for _ in user_info['user_achievements']]
        return user_info

    def _create_new_user(self, member_params):
        return User.objects.create_user(username=member_params.get('email'),
                                        password=const.USER_PASSWORD,
                                        email=member_params.get('email'),
                                        first_name=member_params.get('first_name'),
                                        last_name=member_params.get('last_name'))

    def _get_params_for_member_create(self, params):
        full_name = params.get('full_name')
        separated_full_name = full_name.strip().split(' ', maxsplit=2)
        return {
            'first_name': separated_full_name[1],
            'last_name': separated_full_name[0],
            'father_name': separated_full_name[2] if len(separated_full_name) == 3 else '',
            'phone': convert_phone_format(params['phone']),
            'email': params['email'],
        }


class ExcelFromApps:
    def __init__(self, apps):
        self.apps = apps
        self.wb = Workbook(write_only=True)
        self.sheet = self.wb.create_sheet()

    def add_app_to_sheet(self):
        self._update_column_dimensions()
        header = const.HEADERS_FOR_EXCEL_APP_TABLES
        self.sheet.append(header)
        for app in self.apps:
            row = self._convert_app_to_required_format(app)
            self.sheet.append(row)
        return self._save()

    def _convert_app_to_required_format(self, app):
        birth_day = convert_datetime_to_str(app.birth_day, '%d.%m.%Y')
        draft_season = app.get_draft_time()
        full_name = f"{app.member.user.last_name} {app.member.user.first_name} {app.member.father_name}"
        education_type = [name for ed_type, name in Education.education_program if ed_type == app.education_type]
        education_type = education_type[0] if education_type else ""
        return [full_name, draft_season, birth_day, app.birth_place, app.subject_name, education_type, app.university,
                app.avg_score]

    def _save(self):
        buffer = BytesIO()
        self.wb.save(buffer)
        buffer.seek(0)
        return buffer

    def _update_column_dimensions(self):
        for i in range(1, len(const.HEADERS_FOR_EXCEL_APP_TABLES) + 1):
            letter = get_column_letter(i)
            self.sheet.column_dimensions[letter].width = 20


class Converter:

    @staticmethod
    def convert_education_type_by_name(education_type):
        for program in Education.education_program:
            if program[-1].lower() == education_type.lower():
                return program[0]

    @staticmethod
    def convert_draft_season_by_name(draft_season):
        for s in Application.season:
            if s[-1].lower() == draft_season.lower():
                return s[0]
